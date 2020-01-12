# -*- coding: utf-8 -*-
'''从自己选出的优质股中，读取excel值，查看每个股票的K图，和证券公司的研报，筛选出好的股票存入excel'''
import time
import pandas as pd
from selenium import webdriver
from openpyxl import load_workbook


class ShowK:

    def __init__(self):
        self.share_list0 = []
        self.share_list1 = []

    def read_from_excel(self):
        wb = load_workbook("C:\\Users\\Zhouxiong\\OneDrive\\桌面\\share_of_fund.xlsx")
        ws = wb.get_sheet_by_name('before_check_k')
        first_column = ws['A']
        for x in range(len(first_column)):
            print(first_column[x].value)
            self.share_list0.append(first_column[x].value)
        # return share_list0

    @staticmethod
    def save_to_excel(share_list1):
        wb = load_workbook('C:\\Users\\Zhouxiong\\OneDrive\\桌面\\share_of_fund.xlsx')
        ws = wb.get_sheet_by_name('after_check_k')
        j = ws.max_column
        for z in range(len(share_list1)):
            ws.cell(row=z+1, column=j+1).value = share_list1[z]
        wb.save("C:\\Users\\Zhouxiong\\OneDrive\\桌面\\share_of_fund.xlsx")

    # def save_to_excel():
    #     df1 = pd.DataFrame({
    #         's_name': s_name,
    #         's_code': s_code
    #     })
    #     df1.to_excel('C:/Users/Zhouxiong/OneDrive/桌面/K1.xlsx', sheet_name='after_check')

    def begin(self):
        # s_name = []
        # s_code = []
        # d = pd.read_excel('C:/Users/Zhouxiong/OneDrive/桌面/K0.xlsx', sheet_name='before_check')

        driver = webdriver.Chrome(r'C:\Users\Zhouxiong\AppData\Local\Google\Chrome\Application\chromedriver.exe')  # 打开谷歌浏览器
        num = 0
        self.read_from_excel()
        print(self.share_list0)
        for i in self.share_list0:
            try:
                driver.get('http://quote.cfi.cn/quote_'+str(i)+'.html')
                time.sleep(1)
                driver.find_element_by_link_text(u"日K").click()
                time.sleep(5)
                driver.find_element_by_link_text(u"周K").click()
                time.sleep(5)
                driver.find_element_by_link_text(u"月K").click()
                time.sleep(5)
                driver.find_element_by_link_text(u"研报一览").click()
                time.sleep(5)
                c = input('该股票是否优质,y or n:')
                if c == 'y':
                    num += 1
                    self.share_list1.append(i)
            except:
                print(''+i+'网页获取失败')
                continue
        print('选出的股票数为:'+str(num)+'')
        self.save_to_excel(self.share_list1)
        driver.quit()


if __name__ == '__main__':
    show_k = ShowK()
    show_k.begin()



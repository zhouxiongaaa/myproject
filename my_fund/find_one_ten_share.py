# -*- coding: utf-8 -*-
'''从晨星网获取指定基金的十大持股，并存入excel'''
import time
import json
import requests
from openpyxl import load_workbook
from selenium import webdriver


def enter_page_and_get_json():
    global datas
    url_name = driver.find_element_by_xpath('//*[@id="ctl00_cphMain_gridResult"]/tbody/tr[2]/td[4]/a')
    fund_name = fund_code + url_name.text
    share_list.append(fund_name)
    action = url_name.get_attribute("href")
    url1 = 'http://cn.morningstar.com/handler/quicktake.ashx?command=portfolio&fcid=' + str(action[-10:])
    print(url1)
    wbdata = requests.get(url1).text  # 对HTTP响应的数据JSON化.
    data = json.loads(wbdata)
    # print(data)
    datas = data['Top10StockHoldings']
    return datas


def save_to_excel(share_list):
    wb = load_workbook('C:\\Users\\Zhouxiong\\OneDrive\\桌面\\share_of_fund.xlsx')
    print(wb.get_sheet_names())
    ws = wb.get_sheet_by_name('ten_share')
    j = ws.max_column
    print(j)
    for i in range(len(share_list)):
        ws.cell(row=i+1, column=j+1).value = share_list[i]
    wb.save("C:\\Users\\Zhouxiong\\OneDrive\\桌面\\share_of_fund.xlsx")



if __name__ == '__main__':
    share_list = []
    driver = webdriver.Chrome(r'C:\Users\Zhouxiong\AppData\Local\Google\Chrome\Application\chromedriver.exe')
    driver.get("http://cn.morningstar.com/quickrank/default.aspx")
    time.sleep(3)
    fund_code = input('请输入要获取的基金代码：')
    driver.find_element_by_id("ctl00_cphMain_txtFund").click()
    driver.find_element_by_id("ctl00_cphMain_txtFund").clear()
    driver.find_element_by_id("ctl00_cphMain_txtFund").send_keys(str(fund_code))
    driver.find_element_by_id("ctl00_cphMain_btnGo").click()
    # print(driver.page_source)
    try:
        enter_page_and_get_json()
        for n in range(10):
            try:
                share_list.append(datas[n]['HoldingName'])
            except:
                print('无法获取')
                continue
    except KeyError:
        print('无法获取')

    save_to_excel(share_list)
    print(share_list)
    driver.quit()

